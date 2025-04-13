import tkinter as tk
from tkinter import ttk,filedialog
import os
import openpyxl as op

class PartSearchSettingInstaller():
    def __init__(self):

        self.win = tk.Tk()
        self.win.geometry("500x100")
        self.MakingWindow()
        self.filename = ""
        self.win.mainloop()

    def MakingWindow(self):
        self.win.title("Photonic Tool Installer")
        self.BOX_Frame = ttk.Label(self.win, text='Browse your Part "Tracking Info.xlsx" location')
        self.BOX_Frame.grid(row=0, column=0)
        self.FOLDER_PATH = tk.StringVar()
        # TextLabel = Label()

        AddressFrame = ttk.Labelframe(self.win)
        AddressFrame.grid(row=1,column=0)

        Entry = tk.Entry(AddressFrame, width=70, textvariable=self.FOLDER_PATH)
        Entry.grid(row=0, column=0)

        button2 = tk.Button(AddressFrame,text="Browse", command=self.browse_button)
        button2.grid(row=0, column=1)

        self.CheckResult = tk.StringVar()
        lbl1 = tk.Label(master=self.win, textvariable=self.CheckResult)
        lbl1.grid(row=2, column=0)

        # print(self.FOLDER_PATH)
    def browse_button(self):
        # Allow user to select a directory and store it in global var
        # called folder_path
        self.filename = filedialog.askdirectory()
        self.FOLDER_PATH.set(self.filename)
        self.FileChecker()

    def FileChecker(self, *args):
        # print(self.filename + "/Tracking Info.xlsx")
        if os.path.exists(self.filename + "/Part Number Tracking.xlsx"):
            self.ExcelWriter()
            self.CheckResult.set("Configuration set completed")
        else:
            self.CheckResult.set("This Folder does not contain 'Part Number Tracking.xlsx'")

    def ExcelWriter(self):
        workbook = op.load_workbook("PartSearch\\assets\\PartInfoDatabaseAddress.xlsx")
        worksheet = workbook["Sheet1"]

        max_row = worksheet.max_row + 1
        worksheet.cell(row = max_row , column= 1).value = max_row - 1
        worksheet.cell(row = max_row, column = 2).value = os.environ['COMPUTERNAME']
        worksheet.cell(row = max_row, column = 3).value = "PartInfoDatabaseAddress.xlsx"
        worksheet.cell(row = max_row, column = 4).value = self.filename + "/Part Number Tracking.xlsx"
        workbook.save("PartSearch\\assets\\PartInfoDatabaseAddress.xlsx")

if __name__ == "__main__":
    Folder = PartSearchSettingInstaller()
