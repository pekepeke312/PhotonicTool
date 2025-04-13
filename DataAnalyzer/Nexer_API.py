"""Resources for making Nexar requests."""
import requests
import base64
import json
import time
from typing import Dict
import os, sys
import re
import pandas as pd
import numpy as np
import shutil
import subprocess
from openpyxl import load_workbook
import openpyxl as xl
from openpyxl.styles.borders import Border,Side
from openpyxl.styles.alignment import Alignment

from textwriter import textwriter

NEXAR_URL = "https://api.nexar.com/graphql"
PROD_TOKEN_URL = "https://identity.nexar.com/connect/token"

QUERY_MPN = '''
query Search($mpn: String!) {
    supSearchMpn(
        q: $mpn, 
        currency: "CAD", 
        country: "CAN",
        limit: 1
    ) 
    {
      results {
        part {
          totalAvail
          mpn
          shortDescription
          manufacturer {
            name
          }
          sellers {
          company {
            name
          }
          offers {
            prices {
              #Returns prices for different quantities sellers offer
              #Usually (but not limited to) 1, 10, 100, 1000, 10000 
              quantity
              price
            }
          }
        }      
          specs {
            attribute {
              shortname
            }
            displayValue
          }
        }
      }
    }
    # supSellers{
    # id
    # name
    # homepageUrl
    # }
  }
'''

def get_token(client_id, client_secret):
    """Return the Nexar token from the client_id and client_secret provided."""

    if not client_id or not client_secret:
        raise Exception("client_id and/or client_secret are empty")

    token = {}
    try:
        token = requests.post(
            url=PROD_TOKEN_URL,
            data={
                "grant_type": "client_credentials",
                "client_id": client_id,
                "client_secret": client_secret
            },
            allow_redirects=False,
        ).json()

    except Exception:
        raise

    return token
def decodeJWT(token):
    return json.loads(
        (base64.urlsafe_b64decode(token.split(".")[1] + "==")).decode("utf-8")
    )

class NexarClient:
    def __init__(self, id, secret) -> None:
        self.id = id
        self.secret = secret
        self.s = requests.session()
        self.s.keep_alive = False

        self.token = get_token(id, secret)
        self.s.headers.update({"token": self.token.get('access_token')})
        self.exp = decodeJWT(self.token.get('access_token')).get('exp')

    def check_exp(self):
        if (self.exp < time.time() + 300):
            self.token = get_token(self.id, self.secret)
            self.s.headers.update({"token": self.token.get('access_token')})
            self.exp = decodeJWT(self.token.get('access_token')).get('exp')

    def get_query(self, query: str, variables: Dict) -> dict:
        """Return Nexar response for the query."""
        try:
            self.check_exp()
            r = self.s.post(
                NEXAR_URL,
                json={"query": query, "variables": variables},
            )

        except Exception as e:
            print(e)
            raise Exception("Error while getting Nexar response")

        response = r.json()
        if ("errors" in response):
            for error in response["errors"]: print(error["message"])
            raise SystemExit

        return response["data"]

class Nexer_API():
    def __init__(self, PartNumber):
        self.clientId = os.environ['NEXAR_CLIENT_ID']
        self.clientSecret = os.environ['NEXAR_CLIENT_SECRET']
        self.nexar = NexarClient(self.clientId, self.clientSecret)

        variables = {
            'mpn': PartNumber
        }

        self.results = self.nexar.get_query(QUERY_MPN, variables)

    def getParameter(self, Parameter) -> object:
        specs = self.results.get("supSearchMpn",{}).get("results", {})[0].get("part", {}).get("specs", {})
        if specs:
            results = [i for (i) in specs if i.get('attribute',{}).get('shortname') == Parameter]
            if len(results) > 0:
                return results[0].get('displayValue',{})
        return ''

    def getPartNumber(self):
        if self.results:
            try:
                for it in self.results.get("supSearchMpn", {}).get("results", {}):
                    return it.get("part",{}).get("mpn")
            except:
                return 'no parts found'

    def getManufacturer(self):
        if self.results:
            try:
                for it in self.results.get("supSearchMpn", {}).get("results", {}):
                    return it.get("part",{}).get("manufacturer",{}).get("name", {})
            except:
                return 'no parts found'

    def getDescription(self):
        if self.results:
            try:
                for it in self.results.get("supSearchMpn", {}).get("results", {}):
                    return it.get("part",{}).get("shortDescription")
            except:
                return 'no parts found'

    def getPartType(self):
        if self.results:
            try:
                for it in self.results.get("supSearchMpn", {}).get("results", {}):
                    Description = it.get("part",{}).get("shortDescription")
            except:
                return 'no parts found'

        Type = 'Not Registered'
        try:
            #### Part Type = IC ####
            PartSearch_IC = re.compile(
                r'Transceivers|PMIC|IC|Amplifiers|Amps|Regulators|MCU|EEPROM|Driver|Gates|DAC|Lna|DC-DC|amp|pmic|Comparator|TSSOP|op amp|VSSOP|Dc/Dc|Splitter')
            if type(PartSearch_IC.search(Description)) == re.Match:
                Type = 'IC'

            #### Part Type = Cap ####
            PartSearch_Cap = re.compile(r'Capacitor|Cap')
            if type(PartSearch_Cap.search(Description)) == re.Match:
                Type = 'Capacitor'

            #### Part Type = Attenuator ####
            PartSearch_Cap = re.compile(r'Attenuator')
            if type(PartSearch_Cap.search(Description)) == re.Match:
                Type = 'Chip ATT'

            #### Part Type = Crystal ####
            PartSearch_X = re.compile(r'Crystal|XTAL|Xtal|Resonator|Oscillator|Frequency Control')
            if type(PartSearch_X.search(Description)) == re.Match:
                Type = 'Oscillator'

            #### Part Type = SW ####
            PartSearch_SW = re.compile(r'Switches')
            if type(PartSearch_SW.search(Description)) == re.Match:
                Type = 'SW'

            #### Part Type = LED ####
            PartSearch_LED = re.compile(r'LED|Led')
            if type(PartSearch_LED.search(Description)) == re.Match:
                Type = 'LED'

            #### Part Type = Zener ####
            PartSearch_GDT = re.compile(r'GDT')
            if type(PartSearch_GDT.search(Description)) == re.Match:
                Type = 'GDT'

            #### Part Type = Zener ####
            PartSearch_Zener = re.compile(r'Zener')
            if type(PartSearch_Zener.search(Description)) == re.Match:
                Type = 'Zener'

            #### Part Type = Schottky ####
            PartSearch_Schottky = re.compile(r'Schottky')
            if type(PartSearch_Schottky.search(Description)) == re.Match:
                Type = 'Schottky'

            #### Part Type = TVS ####
            PartSearch_TVS = re.compile(r'TVS')
            if type(PartSearch_TVS.search(Description)) == re.Match:
                Type = 'TVS'

            PartSearch_Cap = re.compile(r'Bridge')
            if type(PartSearch_Cap.search(Description)) == re.Match:
                Type = 'Bridge'

            #### Part Type = Ferrite ####
            PartSearch_Ferrite = re.compile(r'Ferrite|FERRITE')
            if type(PartSearch_Ferrite.search(Description)) == re.Match:
                Type = 'Ferrite'

            #### Part Type = Feed Through ####
            PartSearch_Ferrite = re.compile(r'Feed Through|FEED THROUGH|Feedthrough')
            if type(PartSearch_Ferrite.search(Description)) == re.Match:
                Type = 'Feed Through'

            #### Part Type = BJT ####
            PartSearch_BJT = re.compile(r'BJT')
            if type(PartSearch_BJT.search(Description)) == re.Match:
                Type = 'BJT'

            #### Part Type = MOSFET ####
            PartSearch_MOSFET = re.compile(r'MOSFET')
            if type(PartSearch_MOSFET.search(Description)) == re.Match:
                Type = 'MOSFET'

            #### Part Type = FUSE ####
            PartSearch_FUSE = re.compile(r'FUSE')
            if type(PartSearch_FUSE.search(Description)) == re.Match:
                Type = 'FUSE'

            #### Part Type = Resistors ####
            PartSearch_Res = re.compile(r'Resistor|Res|RESISTOR')
            if type(PartSearch_Res.search(Description)) == re.Match:
                Type = 'Resistor'

            #### Part Type = Trimmer ####
            PartSearch_Res = re.compile(r'TRIMMER|Trimmer')
            if type(PartSearch_Res.search(Description)) == re.Match:
                Type = 'Trimmer'

            #### Part Type = Thermistor ####
            PartSearch_Res = re.compile(r'Thermistor')
            if type(PartSearch_Res.search(Description)) == re.Match:
                Type = 'Thermistor'

            #### Part Type = Inductors ####
            PartSearch_Res = re.compile(r'IND|Inductor|Coil|Power Choke')
            if type(PartSearch_Res.search(Description)) == re.Match:
                Type = 'Inductor'

            #### Part Type = Connectors ####
            PartSearch_Res = re.compile(r'Connector|Header|Terminal|Adapter|Plug|TERMINAL|CONN|Conn|Receptacle|Pomona Electronics 5085|Jack')
            if type(PartSearch_Res.search(Description)) == re.Match:
                Type = 'Connector'

            #### Part Type = Switch ####
            PartSearch_Res = re.compile(r'Switch|SPDT')
            if type(PartSearch_Res.search(Description)) == re.Match:
                Type = 'Switch'

        except:
            pass

        return Type


    def getLifecycle(self):
        return self.getParameter(Parameter='lifecyclestatus')

    def getValue(self):
        inductance = self.getParameter(Parameter="inductance")
        resistance = self.getParameter(Parameter="resistance")
        capacitance = self.getParameter(Parameter="capacitance")
        return inductance + ' ' + capacitance + ' ' + resistance

    def getStock(self):
        if self.results:
            try:
                for it in self.results.get("supSearchMpn", {}).get("results", {}):
                    return f'{it.get("part", {}).get("totalAvail",{}):,}' # int -> Digit Grouping format
            except:
                return 'no parts found'

    def getRohsStatus(self):
        return self.getParameter(Parameter='rohs')

    def getTolerance(self):
        return self.getParameter(Parameter='tolerance')

    def getPackage(self):
        return self.getParameter(Parameter='case_package')

    def getDielectric(self):
        return self.getParameter(Parameter='dielectric')

    def getMount(self):
        return self.getParameter(Parameter='termination')

    def getVoltageRating(self):
        return self.getParameter(Parameter='voltagerating')

    def getWattageRating(self):
        return self.getParameter(Parameter='powerrating')

    def getTemperatureCoefficient(self):
        return self.getParameter(Parameter='temperaturecoefficient')

    def getMaterial(self):
        return self.getParameter(Parameter='composition')

    def getLength(self):
        return self.getParameter(Parameter='length')

    def getHeight(self):
        return self.getParameter(Parameter='height')

    def getWidth(self):
        return self.getParameter(Parameter='width')
    def getOperatingtemp(self):
        min = self.getParameter(Parameter='minoperatingtemperature')
        max = self.getParameter(Parameter='maxoperatingtemperature')
        return f'{min} \u301C {max}'

    def getOperatingtemp_min(self):
        return self.getParameter(Parameter='minoperatingtemperature')

    def getOperatingtemp_max(self):
        return self.getParameter(Parameter='maxoperatingtemperature')

    def getAltiumDescription(self):
        Type=''
        if self.getPartType() == "Resistor":
            Description = 'RES ' + self.getValue().replace(" ", "") + ' ' + self.getWattageRating().replace(" ", "") + ' ' + self.getPackage()
            return Description

        elif self.getPartType() == "Capacitor":
            Description = ('CAP '
                           + self.getValue().replace(" ", "")
                           + ' ' + self.getTolerance().replace(" ", "")
                           + ' ' + self.getVoltageRating().replace(" ", "")
                           + ' ' + self.getPackage())
            return Description

        elif self.getPartType() == "Inductor":
            Value = self.getParameter(Parameter="inductance").replace(" ", "")
            Currentrating = self.getParameter(Parameter="currentrating").replace(" ", "")
            Description = 'IND ' + Value + ' ' + self.getTolerance().replace(" ", "") + ' ' + Currentrating + ' ' + self.getPackage()
            return Description

        elif self.getPartType() == "Connector":
            Description = 'Not Supporting Yet'
            return Description

        elif (self.getPartType() == "TVS"
              or self.getPartType() == "Zener"
              or self.getPartType() == "Schottky"
              or self.getPartType() == "Bridge"
        ):
            Description = self.getDescription()
            Zenorvoltage = self.getParameter(Parameter="reversebreakdownvoltage").replace(" ", "")
            PartSearch_Cap = re.compile(r'Zener')
            if type(PartSearch_Cap.search(Description)) == re.Match:
                Type = 'ZENER'

            PartSearch_Cap = re.compile(r'TVS')
            if type(PartSearch_Cap.search(Description)) == re.Match:
                Type = 'TVS'

            PartSearch_Cap = re.compile(r'Schottky')
            if type(PartSearch_Cap.search(Description)) == re.Match:
                Type = 'SCHOTTKEY'

            PartSearch_Cap = re.compile(r'Bridge')
            if type(PartSearch_Cap.search(Description)) == re.Match:
                Type = 'BRIDGE'

            Description = 'DIODE ' + Type + ' ' + self.getPackage() + ' ' + Zenorvoltage
            return Description

        elif self.getPartType() == "LED":
            Color = self.getParameter(Parameter="illuminationcolor").replace(" ", "")
            Curren = self.getParameter(Parameter="forwardcurrent").replace(" ", "")
            Voltage = self.getParameter(Parameter="forwardvoltage").replace(" ", "")
            Description = 'LED ' + Color + ' ' + Curren + ' ' + Voltage + ' ' + self.getPackage()
            return Description

        elif self.getPartType() == 'Oscillator':
            Description = self.getDescription()
            PartSearch_Cap = re.compile(r'Tcxo')
            if type(PartSearch_Cap.search(Description)) == re.Match:
                Type = 'TCXO'

            PartSearch_Cap = re.compile(r'Xo')
            if type(PartSearch_Cap.search(Description)) == re.Match:
                Type = 'XO'

            PartSearch_Cap = re.compile(r'VCO|Vco')
            if type(PartSearch_Cap.search(Description)) == re.Match:
                Type = 'VCO'

            PartSearch_Cap = re.compile(r'VCXO|Vcxo')
            if type(PartSearch_Cap.search(Description)) == re.Match:
                Type = 'VCXO'

            Description = 'OSC ' + Type + 'Not Supporting Yet'
            return Description

class Reader_from_Nexer:
    def __init__(self, PATH=""):
        if PATH != "":
            self.BOMReader(PATH)
            self.Get_Multiple_Data_From_Nexer(PNList=list(self.PartNumberList_BOM.keys()))
            #self.ColumnOrderUpdate()
            self.WritingExcel()
    def BOMReader(self, PATH):
        starttime = time.time()
        try:
            self.BOM_data = pd.read_excel(PATH)
        except:
            self.BOM_data = pd.read_csv(PATH) #back up for the case file type is CSV

        try:
            ItemID_List = np.array(self.BOM_data['item number'])
        except:
            try:
                ItemID_List = np.array(self.BOM_data['7-digit PN'])
            except:
                try:
                    ItemID_List = np.array(self.BOM_data['Part Number'])
                except:
                    pass
                pass

        try:
            Mfr_List = np.array(self.BOM_data['Vendor Parts'])
        except:
            try:
                Mfr_List = np.array(self.BOM_data['Manufacturer Part Number'])
            except:
                try:
                    Mfr_List = np.array(self.BOM_data['Manufacturer PN'])
                except:
                    try:
                        Mfr_List = np.array(self.BOM_data['manufacturer item number_1'])
                    except:
                        pass
                    pass
                pass

        self.PartNumberList_BOM = {}
        for Index, Value in enumerate(Mfr_List):
                #        for n in range(len(Mfr_List)):
            if type(Value) == str:
                for PN_temp in Value.split(', '):
                    idx = PN_temp.find('(')
                    if idx == -1:
                       PN = PN_temp
                    else:
                        PN = PN_temp[:idx]
                    self.PartNumberList_BOM[PN] = Value

        self.PhotonicPNList_BOM = {}
        for Index, Value in enumerate(ItemID_List):
            self.PhotonicPNList_BOM[Mfr_List[Index]] = Value

        elapstedtime = time.time() - starttime

        print("BOM Checked in {:.3}s".format(elapstedtime))
        textwriter("BOM Checked in {:.3}s".format(elapstedtime))

    def Get_Multiple_Data_From_Nexer(self, PNList=[""]):
        Result_DF = pd.DataFrame()
        for PNn in range(len(list(PNList))):
            starttime = time.time()
            self.PartInfo = {}
            PartData = Nexer_API(PartNumber=PNList[PNn])

            self.indexlist = [
                ("Photonic#", ""),
                ("Manufacturer Part Number", ""), #0
                ("Type", ""), #1
                ("Value", ""), #2
                ("Tolerance", ""), #3
                ("Package", ""), #4
                ("Size", "Length"),  # 5
                ("Size", "Width"),  # 6
                ("Size", "Height"),  # 7
                ("Rating", "Voltage"), #8
                ("Rating","Power"), #9
                ("Dielectric", ""), #10
                ("Operating Temp","Min"), #11
                ("Operating Temp", "Max"),  # 12
                ("Manufacturer", ""), #13
                ("Lifecycle Status", ""), #14
                ("Stock", "Nexer"), #15
                ("Compliance", "RoHS"), #16
                ("Description", "Nexer-API"), #17
                ("Description", "For Altium") #18
            ]

            Datadict = {}
            ### Photonic PN ###
            ## Not Yet
            Datadict[self.indexlist[0]] = self.PhotonicPNList_BOM[PNList[PNn]]
            Datadict[self.indexlist[1]] = PNList[PNn] ### Manufacturer Part Number
            Datadict[self.indexlist[2]] = PartData.getPartType() ### Type
            Datadict[self.indexlist[3]] = PartData.getValue()  ### Value
            Datadict[self.indexlist[4]] = PartData.getTolerance()  ### Tolerance
            Datadict[self.indexlist[5]] = PartData.getPackage()  ### Package
            Datadict[self.indexlist[6]] = PartData.getLength()  ### Length
            Datadict[self.indexlist[7]] = PartData.getWidth()  ### Width
            Datadict[self.indexlist[8]] = PartData.getHeight()  ### Height
            Datadict[self.indexlist[9]] = PartData.getVoltageRating()  ### Rating - Voltage
            Datadict[self.indexlist[10]] = PartData.getWattageRating()  ### Rating - Wattage
            Datadict[self.indexlist[11]] = PartData.getDielectric()  ### Dielectric
            Datadict[self.indexlist[12]] = PartData.getOperatingtemp_min()  ### Temperature Range Min
            Datadict[self.indexlist[13]] = PartData.getOperatingtemp_max()  ### Temperature Range Max
            Datadict[self.indexlist[14]] = PartData.getManufacturer()  ### Manufacturer
            Datadict[self.indexlist[15]] = PartData.getLifecycle()  ### Lifecycle
            Datadict[self.indexlist[16]] = PartData.getStock()  ### Stock Availability
            Datadict[self.indexlist[17]] = PartData.getRohsStatus()  ### RoHS Status
            Datadict[self.indexlist[18]] = PartData.getDescription()  ### Description
            Datadict[self.indexlist[19]] = PartData.getAltiumDescription()  ### Altium Description

            #### Index Update ####
            cols = pd.MultiIndex.from_tuples(self.indexlist)
            Temp_DataFrame = pd.DataFrame(columns=cols)
            Temp_DataFrame.loc[list(PNList)[PNn]] = Datadict
            Result_DF = pd.concat([Result_DF, Temp_DataFrame])

            elapstedtime = time.time() - starttime
            print("No.{}/{}: Data of {} Derived in {:.3}s".format(PNn + 1, len(list(PNList)), list(PNList)[PNn],
                                                                  elapstedtime))
            textwriter("No.{}/{}: Data of {} Derived in {:.3}s".format(PNn + 1, len(list(PNList)), list(PNList)[PNn],
                                                                       elapstedtime))
        self.Result_DataFrame = Result_DF

    def AutoExcelWidthAdjuster(self, path="FilePath"):
        # read input xlsx
        wb1 = xl.load_workbook(filename=path)
        ws1 = wb1.worksheets[0]

        # set column width
        for col in ws1.columns:
            max_length = 0
            column = col[0].column

            for cell in col:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))

            adjusted_width = (max_length + 2)
            ws1.column_dimensions[col[1].column_letter].width = adjusted_width

        # border setting
        border = Border(top=Side(style='dotted', color='000000'),
                        bottom=Side(style='dotted', color='000000'),
                        left=Side(style='thin', color='000000'),
                        right=Side(style='thin', color='000000')
                        )

        for row_num in range(1, ws1.max_row+1):
            for col_num in range(1, ws1.max_column+1):
                ws1.cell(row=row_num, column=col_num).border = border
                ws1.cell(row=row_num, column=col_num).alignment = Alignment(horizontal='center',
                                                                            vertical='center',
                                                                            wrap_text=False)


        ws1.cell(row=1, column=1).value = 'Index'
        ws1.views.activeCell = 'A3'
        # save xlsx file
        wb1.save(path)
    def WritingExcel(self):
        print("\nMaking Result Sheet at Desktop...")
        textwriter("\nMaking Result Sheet at Desktop...")

        try:
            path = os.environ['DataAnalyzerPath'] + str('\\AnalysisTool\\DataAnalyzer\\assets\\')
        except:
            path = os.getcwd() + str('\\DataAnalyzer\\assets\\')  # For normal operation
            if __name__ == '__main__':
                path = os.getcwd() + str('\\assets\\')  # For normal operation

            # if __name__ == "__main__":
            #     path = os.getcwd() + str('\\assets\\')      #For debuging case
            # else:
            #     path = os.getcwd() + str('\\DataAnalyzer\\assets\\')   #For normal operation
        path = r'{}'.format(path)

        path_desktop = os.environ['USERPROFILE'] + str('\\Desktop\\')
        try:
            os.system('TASKKILL /F /IM excel.exe')
        except:
            pass
        ### Delete old file
        try:
            os.remove(path_desktop + "DerivedExcel.xlsx")
        except:
            pass

        ### Copy all excel format and conditional format settings
        shutil.copyfile(path + 'Base.xlsx', path_desktop + 'DerivedExcel.xlsx')


        book = load_workbook(path_desktop + 'DerivedExcel.xlsx')
        writer = pd.ExcelWriter(path_desktop + 'DerivedExcel.xlsx', engine='openpyxl')
        writer.book = book

        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

        self.Result_DataFrame.to_excel(writer, "Sheet1",)
        writer.save()
        writer.close()

        self.AutoExcelWidthAdjuster(path_desktop + 'DerivedExcel.xlsx')
        # self.DigikeyPartInfo.terminate()

        subprocess.Popen(['start', path_desktop + 'DerivedExcel.xlsx'], shell=True)

if __name__ == '__main__':
    InputPartNuber = 'HMC998APM5E'
    API = Nexer_API(PartNumber=InputPartNuber)
    print(f'Input Part Number : {InputPartNuber}')
    print(f'Part Number : {API.getPartNumber()}')
    print(f'Part Type : {API.getPartType()}')
    print(f'Manufacturer : {API.getManufacturer()}')
    print(f'Description : {API.getDescription()}')
    print(f'Value : {API.getValue()}')
    print(f'Lifecycle : {API.getLifecycle()}')
    print(f'Stock : {API.getStock()}')
    print(f'RoHS: {API.getRohsStatus()}')
    print(f'Operating Temperature: {API.getOperatingtemp()}')
    print(f'Tolerance: {API.getTolerance()}')
    print(f'Package: {API.getPackage()}')
    print(f'Dielectric: {API.getDielectric()}')
    print(f'Mount: {API.getMount()}')
    print(f'Voltage Rating: {API.getVoltageRating()}')
    print(f'Wattage Rating: {API.getWattageRating()}')
    print(f'Temperature Coefficient: {API.getTemperatureCoefficient()}')
    print(f'Material: {API.getMaterial()}')
    print(f'Desicription: {API.getAltiumDescription()}')
    print()

    # Path = r"C:\Users\IsaoYoneda\Desktop\BOM_Sample.xlsx"
    # Run = Reader_from_Nexer(Path)