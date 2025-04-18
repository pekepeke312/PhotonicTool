import numpy as np
from openpyxl import load_workbook
import openpyxl as xl
from openpyxl.styles.borders import Border,Side
from openpyxl.styles.alignment import Alignment
import requests
import json
import re
import pandas as pd
import time
import subprocess
import shutil
import os
import pathlib

SCRAPING_TIME_MARGIN = 5

from textwriter import textwriter
if __name__ != "__main__":
    from .Digikey_WebScraping import Digikey_WebScraping
else:
    from Digikey_WebScraping import Digikey_WebScraping

PartSearchPath = str(pathlib.Path(__file__).parent.parent)
PartInfoDatabaseAddress = PartSearchPath + str('\\PartSearch\\assets\\PartInfoDatabaseAddress.xlsx')

from PartSearch.PartSearchTool import PartSearchTool#, PartInfoDatabaseAddress


class API_Mouser():
    def __init__(self, PNList = [""], PN="",PATH="", APIKEY="", APIver="v1.0"):
        self.URL_Endpoint = 'https://api.mouser.com/api/' + APIver + '/search/keyword'
        self.API_Key = APIKEY
        self.TargetPartNumber = PN
        self.PartList = PNList
        self.PartNumberList_BOM = {}
        self.InternalDatabaseLoader()
        self.DigikeyPartInfo = Digikey_WebScraping()

        ### Methods ###
        ### Single Part Number Case
        if PN != "":
            self.Get_From_Mouser(self.TargetPartNumber)
            # self.InternalDatabaseLoader()
            self.ParameterUpdate()
        ### Multipule Number Case
        elif PNList != [""]:
            self.Get_Multiple_Data_From_Mouser(PNList=self.PartList)
            # self.InternalDatabaseLoader()
            self.ColumnOrderUpdate()
            self.WritingExcel()
        elif PATH !="":
            self.BOMReader(PATH)
            self.Get_Multiple_Data_From_Mouser(PNList = list(self.PartNumberList_BOM.keys()))
            # self.InternalDatabaseLoader()
            self.ColumnOrderUpdate()
            self.WritingExcel()
        print("")

    def BOMReader(self, PATH):
        starttime = time.time()
        self.BOM_data = pd.read_excel(PATH)

        try:
            ItemID_List = np.array(self.BOM_data['Item ID'])
        except:
            try:
                ItemID_List = np.array(self.BOM_data['7-digit PN'])
            except:
                try:
                    ItemID_List = np.array(self.BOM_data['Part Number'])
                except:
                    pass
                pass

        try:
            Mfr_List = np.array(self.BOM_data['Vendor Parts'])
        except:
            try:
                Mfr_List = np.array(self.BOM_data['Manufacturer Part Number'])
            except:
                try:
                    Mfr_List = np.array(self.BOM_data['Manufacturer PN'])
                except:
                    try:
                        Mfr_List = np.array(self.BOM_data['Manufacturer Part Number 1'])
                    except:
                        pass
                    pass
                pass

        #ItemID_List = np.array(self.BOM_data['Item ID'])
        #Mfr_List = np.array(self.BOM_data['Vendor Parts'])
        #self.ItemID_Number = len(Mfr_List)
        self.PartNumberList_BOM = {}
        for Index, Value in enumerate(Mfr_List):
#        for n in range(len(Mfr_List)):
            if type(Value) == str:
                for PN_temp in Value.split(', '):
                    idx = PN_temp.find('(')
                    if idx == -1:
                        PN= PN_temp
                    else:
                        PN = PN_temp[:idx]
                    self.PartNumberList_BOM[PN] = Value

        elapstedtime = time.time() - starttime

        print("BOM Checked in {:.3}s".format(elapstedtime))
        textwriter("BOM Checked in {:.3}s".format(elapstedtime))

        if elapstedtime < SCRAPING_TIME_MARGIN:
            Additonal_Sleep_Time = SCRAPING_TIME_MARGIN - elapstedtime
            time.sleep(Additonal_Sleep_Time)
            print("Operation was too fast; Sleeping for {:.3}s".format(Additonal_Sleep_Time))
            textwriter("Operation was too fast; Sleeping for {:.3}s".format(Additonal_Sleep_Time))

    def InternalDatabaseLoader(self):
        starttime = time.time()
        text = "--- Loading Photonic Part Numbers from Perforce Database"
        print(text)
        textwriter(text)

        PartSearch = PartSearchTool()
        InternalDatabase = PartSearch.DatabaseLoader(Path=PartInfoDatabaseAddress)
        ExcelRaw = pd.read_excel(InternalDatabase["Address"], sheet_name=None, header=None)

        ## Finding Valid Sheet Names ------------------
        L_Valid_Sheet_Names = []
        for Value in ExcelRaw.keys():
            if (re.search(r'\d{2}', Value)):
                L_Valid_Sheet_Names.append(Value)

        ## Making Database -----------------
        self.PartNumberDatabase = []

        for Sheet in L_Valid_Sheet_Names:
            for n in range(len(ExcelRaw[Sheet].values)):
                for k in range(len(ExcelRaw[Sheet].values[n, :])):
                    try:
                        if (ExcelRaw[Sheet].values[n, k] == "Mfg1 PN") or (
                                ExcelRaw[Sheet].values[n, k] == "Manufacturer PN"):
                            Column_ManufacturerPN = k
                    except:
                        Column_ManufacturerPN = 'nan'

                    if ExcelRaw[Sheet].values[n, k] == ("7-digit PN"):
                        Column_PartNumber = k

            for n in range(1, len(ExcelRaw[Sheet].values)):
                try:
                    PartNumber = ExcelRaw[Sheet].values[n, Column_PartNumber]
                except:
                    print()

                try:
                    ManufacturerPN = ExcelRaw[Sheet].values[n, Column_ManufacturerPN]
                except:
                    ManufacturerPN = ""

                self.PartNumberDatabase.append(
                    (
                        PartNumber,
                        ManufacturerPN,
                    )
                )
        self.DF_PartDataBase = pd.DataFrame(self.PartNumberDatabase, dtype='str')
        self.DF_PartDataBase.columns = ["PartNumber",
                                        "Manufacturer PN",
                                        ]

        elapstedtime = time.time() - starttime
        text = 'Loading the all Parts in {:.3}s'.format(elapstedtime)
        print(text)
        textwriter(text)

    def Get_Photonic_PartNumber(self, MfrPN = ""):
        Photonic_PN = str(self.DF_PartDataBase[self.DF_PartDataBase["Manufacturer PN"] == MfrPN]["PartNumber"].values[0])
        return Photonic_PN

    def Get_From_Mouser(self, PartNumber = ""):
        headers = {
            'accept': 'text/json',
            'Content-Type': 'application/json',
        }

        params = (
            ('apiKey', self.API_Key),
        )

        data = str(
            {
                "SearchByKeywordRequest": {
                    "keyword": PartNumber,
                    "searchOptions": "string",
                    "searchWithYourSignUpLanguage": "string"
                }
            }
        )
        response = requests.post(self.URL_Endpoint, headers=headers, params=params, data=data)

        if response.status_code == 429 or response.status_code == 403:
            print("too many request! Need to wait")
            textwriter("too many request! Need to wait")
            self.PartInfo = {}

        elif response.status_code == 200:
            response_content_json = json.loads(response.content.decode('utf-8'))
            if len(response_content_json['SearchResults']['Parts']) > 0:
                self.PartInfo = response_content_json['SearchResults']['Parts'][0]
            else:
                self.PartInfo = {}

            ### Deleting wrong information to avoid misleading
            if self.PartParameter("ManufacturerPartNumber") != PartNumber:
                self.PartInfo = {}

    def PartParameter(self,info=""):
        try:
            return self.PartInfo[info]
        except:
            self.PartInfo[info] = '-'
            return self.PartInfo[info]

    def Get_Multiple_Data_From_Mouser(self, PNList = [""]):
        Result_DF = pd.DataFrame()
        for PNn in range(len(list(PNList))):
            # if PNn%1 == 0:
            #     print("Break Time for API System in 1.5sec...")
            #     textwriter("Break Time for API System in 1.5sec...")
            #     time.sleep(1.5)
            starttime = time.time()
            self.PartInfo = {}
            self.DigikeyStock = ''
            self.Get_From_Mouser(list(PNList)[PNn])
            self.DigikeyPartInfo.TargetPN = list(PNList)[PNn]
            self.DigikeyStock = self.DigikeyPartInfo.Scraping()
            self.ParameterUpdate()

            self.indexlist = [
                ("Photonic#", ""),
                ("Part Number", ""),
                ("Category", ""),
                ("Type", ""),
                ("Value", ""),
                ("Tolerance", ""),
                ("Size", ""),
                ("Rating", ""),
                ("Dielectric", ""),
                ("Operating Temp",""),
                ("Manufacturer", ""),
                ("Lifecycle Status", ""),
                ("Lead Time", ""),
                ("Stock", "Mouser"),
                ("Stock", "Digi-Key"),
            ]

            Datadict ={}
            try:
                # Datadict[self.indexlist[0]] = self.PartNumberList_BOM[list(PNList)[PNn]]
                Datadict[self.indexlist[0]] = self.Get_Photonic_PartNumber(self.PartParameter("ManufacturerPartNumber"))
            except:
                pass
            Datadict[self.indexlist[1]] = self.PartParameter("ManufacturerPartNumber")
            Datadict[self.indexlist[2]] = self.PartParameter("Category")
            Datadict[self.indexlist[3]] = self.PartParameter("Type")
            Datadict[self.indexlist[4]] = self.PartParameter("Value")
            Datadict[self.indexlist[5]] = self.PartParameter("Tolerance")
            Datadict[self.indexlist[6]] = self.PartParameter("Size")
            Datadict[self.indexlist[7]] = self.PartParameter("Rating")
            Datadict[self.indexlist[8]] = self.PartParameter("Dielectric")
            Datadict[self.indexlist[9]] = self.PartParameter("Temperature")
            Datadict[self.indexlist[10]] = self.PartParameter("Manufacturer")
            Datadict[self.indexlist[11]] = self.PartParameter("LifecycleStatus")
            Datadict[self.indexlist[12]] = self.PartParameter("LeadTime")
            Datadict[self.indexlist[13]] = self.PartParameter("Availability")
            Datadict[self.indexlist[14]] = self.DigikeyStock

            try:
                for n in range(len(self.PartInfo["PriceBreaks"])):
                    Datadict["Mouser Price [CAD]","Qty:"+ str(self.PartInfo["PriceBreaks"][n]["Quantity"])] = self.PartInfo["PriceBreaks"][n]['Price']
                    self.indexlist.append(("Mouser Price [CAD]","Qty:"+str(self.PartInfo["PriceBreaks"][n]["Quantity"])))

                Datadict["Compliance","RoHS"] = self.PartParameter("ROHSStatus")
                self.indexlist.append(("Compliance","RoHS"))
                for n in range(len(self.PartInfo["ProductCompliance"])):
                    Datadict["Compliance",self.PartInfo["ProductCompliance"][n]["ComplianceName"]] = \
                    self.PartInfo["ProductCompliance"][n]["ComplianceValue"]
                    self.indexlist.append(("Compliance", str(self.PartInfo["ProductCompliance"][n]["ComplianceName"])))
            except:
                pass

            Datadict["Data Sheet",""] = self.PartParameter("DataSheetUrl")
            Datadict["Part Image",""] = self.PartParameter("ImagePath")

            self.indexlist.append(("Data Sheet",""))
            self.indexlist.append(("Part Image",""))


            #### Index Update ####
            cols = pd.MultiIndex.from_tuples(self.indexlist)
            Temp_DataFrame = pd.DataFrame(columns=cols)
            Temp_DataFrame.loc[list(PNList)[PNn]] = Datadict
            Result_DF = pd.concat([Result_DF,Temp_DataFrame])

            elapstedtime = time.time() - starttime
            print("No.{}/{}: Data of {} Derived in {:.3}s".format(PNn+1,len(list(PNList)), list(PNList)[PNn], elapstedtime))
            textwriter("No.{}/{}: Data of {} Derived in {:.3}s".format(PNn+1,len(list(PNList)), list(PNList)[PNn], elapstedtime))
        self.Result_DataFrame = Result_DF

    def ParameterUpdate(self):
        #starttime = time.time()
        #### Part Type = Res ####
        try:
            #### Part Type = Crystal ####
            PartSearch_X=re.compile(r'Crystal|XTAL|Crystal|Resonator')
            if type(PartSearch_X.search(self.PartInfo["Description"])) == re.Match:
                self.PartInfo['Type'] = 'Oscillator'

            #### Part Type = SW ####
            PartSearch_SW=re.compile(r'Switches')
            if type(PartSearch_SW.search(self.PartInfo["Description"])) == re.Match:
                self.PartInfo['Type'] = 'SW'

            #### Part Type = LED ####
            PartSearch_LED=re.compile(r'LED')
            if type(PartSearch_LED.search(self.PartInfo["Description"])) == re.Match:
                self.PartInfo['Type'] = 'LED'

            #### Part Type = Zener ####
            PartSearch_GDT=re.compile(r'GDT')
            if type(PartSearch_GDT.search(self.PartInfo["Description"])) == re.Match:
                self.PartInfo['Type'] = 'GDT'

            #### Part Type = Zener ####
            PartSearch_Zener=re.compile(r'Zener')
            if type(PartSearch_Zener.search(self.PartInfo["Description"])) == re.Match:
                self.PartInfo['Type'] = 'Zener'

            #### Part Type = Schottky ####
            PartSearch_Schottky=re.compile(r'Schottky')
            if type(PartSearch_Schottky.search(self.PartInfo["Description"])) == re.Match:
                self.PartInfo['Type'] = 'Schottky'

            #### Part Type = TVS ####
            PartSearch_TVS=re.compile(r'TVS')
            if type(PartSearch_TVS.search(self.PartInfo["Description"])) == re.Match:
                self.PartInfo['Type'] = 'TVS'

            #### Part Type = Ferrite ####
            PartSearch_Ferrite=re.compile(r'Ferrite|FERRITE')
            if type(PartSearch_Ferrite.search(self.PartInfo["Description"])) == re.Match:
                self.PartInfo['Type'] = 'Ferrite'

            #### Part Type = Feed Through ####
            PartSearch_Ferrite=re.compile(r'Feed Through|FEED THROUGH|Feedthrough')
            if type(PartSearch_Ferrite.search(self.PartInfo["Description"])) == re.Match:
                self.PartInfo['Type'] = 'Feed Through'

            #### Part Type = BJT ####
            PartSearch_BJT=re.compile(r'BJT')
            if type(PartSearch_BJT.search(self.PartInfo["Description"])) == re.Match:
                self.PartInfo['Type'] = 'BJT'

            #### Part Type = MOSFET ####
            PartSearch_MOSFET = re.compile(r'MOSFET')
            if type(PartSearch_MOSFET.search(self.PartInfo["Description"])) == re.Match:
                self.PartInfo['Type'] = 'MOSFET'

            #### Part Type = FUSE ####
            PartSearch_FUSE = re.compile(r'FUSE')
            if type(PartSearch_FUSE.search(self.PartInfo["Description"])) == re.Match:
                self.PartInfo['Type'] = 'FUSE'

            #### Part Type = Resistors ####
            PartSearch_Res = re.compile(r'Resistors')
            if type(PartSearch_Res.search(self.PartInfo["Description"])) == re.Match:
                self.PartInfo['Type'] = 'Resistor'

            #### Part Type = Trimmer ####
            PartSearch_Res = re.compile(r'TRIMMER|Trimmer')
            if type(PartSearch_Res.search(self.PartInfo["Description"])) == re.Match:
                self.PartInfo['Type'] = 'Trimmer'

            #### Part Type = Thermistor ####
            PartSearch_Res = re.compile(r'Thermistor')
            if type(PartSearch_Res.search(self.PartInfo["Description"])) == re.Match:
                self.PartInfo['Type'] = 'Thermistor'

            #### Part Type = IC ####
            PartSearch_IC = re.compile(r'Transceivers|PMIC|IC|Amplifiers|Amps|Regulators|MCU|EEPROM|Driver|Gates|DAC')
            if type(PartSearch_IC.search(self.PartInfo["Description"])) == re.Match:
                self.PartInfo['Type'] = 'IC'

            #### Part Type = Inductors ####
            PartSearch_Res = re.compile(r'IND|Inductor|Coil')
            if type(PartSearch_Res.search(self.PartInfo["Description"])) == re.Match:
                self.PartInfo['Type'] = 'Inductor'

            #### Part Type = Connectors ####
            PartSearch_Res = re.compile(r'Connector|Header|Terminal|Adapter|Plug|TERMINAL|CONN|Conn|Receptacle')
            if type(PartSearch_Res.search(self.PartInfo["Description"])) == re.Match:
                self.PartInfo['Type'] = 'Connector'

            #### Part Type = Cap ####
            PartSearch_Cap=re.compile(r'Capacitors')
            if type(PartSearch_Cap.search(self.PartInfo["Description"])) == re.Match:
                self.PartInfo['Type'] = 'Capacitor'

        except:
            pass

        try:
            if self.PartInfo['Type'] == 'Resistor':
                self.PartInfo['Value'] = self.DigikeyPartInfo.PartInfo['Resistance']

                #### Wattage Rating ####
                self.PartInfo['Rating'] = self.DigikeyPartInfo.PartInfo['Power (Watts)']

                #self.PartInfo['Dielectric'] = self.DigikeyPartInfo.PartInfo['Operating Temperature']


        except:
            pass

        try:
            if self.PartInfo['Type'] == 'Ferrite':
                #idx = self.DigikeyPartInfo.PartInfo['Impedance'].find('@ Frequency ')
                self.PartInfo['Value'] = self.DigikeyPartInfo.PartInfo['Impedance @ Frequency']#[idx+len('@ Frequency '):]

                #idx = self.DigikeyPartInfo.PartInfo['Current'].find('Rating ')
                self.PartInfo['Rating'] = self.DigikeyPartInfo.PartInfo['Current Rating (Max)']#[idx+len('Rating '):]
                self.PartInfo['Tolerance'] = self.DigikeyPartInfo.PartInfo['DC Resistance (DCR) (Max)'] + ' MAX'
        except:
            pass

        try:
            if self.PartInfo['Type'] == 'TVS':
                self.PartInfo['Value'] = self.DigikeyPartInfo.PartInfo['Voltage']
                self.PartInfo['Rating'] = self.DigikeyPartInfo.PartInfo['Current']
            if self.PartInfo['Type'] == 'Zener':
                self.PartInfo['Rating'] = self.DigikeyPartInfo.PartInfo['Power']
                self.PartInfo['Value'] = self.DigikeyPartInfo.PartInfo['Voltage']
        except:
            pass

        try:
            if self.PartInfo['Type'] == 'IC':
                self.PartInfo['Rating'] = self.DigikeyPartInfo.PartInfo['Voltage - Supply']
                self.PartInfo['Value'] = self.DigikeyPartInfo.PartInfo['Protocol']
        except:
            pass

        try:
            if self.PartInfo['Type'] == 'Capacitor':

                #### Capacitance ####
                Search_Value_F = re.compile(r'(\d+(?:\.\d+)?)(p|u)?F')
                if type(Search_Value_F.search(self.PartInfo["Description"])) == re.Match:
                    self.PartInfo['Value'] = str(Search_Value_F.search(self.PartInfo["Description"])[0])

                #### Dielectric ####
                self.PartInfo['Dielectric'] = self.DigikeyPartInfo.PartInfo['Temperature Coefficient']

                #### Voltage Rating ####
                self.PartInfo['Rating'] = self.DigikeyPartInfo.PartInfo['Voltage - Rated']
        except:
            pass

        try:
            #### Tolerance ####
            self.PartInfo['Tolerance'] = self.DigikeyPartInfo.PartInfo['Tolerance']
            #Search_Tolerance = re.compile(r'\d+( )?%')
            # if type(Search_Tolerance.search(self.PartInfo["Description"])) == re.Match:
            #     self.PartInfo['Tolerance'] = str(Search_Tolerance.search(self.PartInfo["Description"])[0])
        except:
            pass

        try:
            #### Case Code ####
            #idx = self.DigikeyPartInfo.PartInfo['Package'].find('Case ')
            self.PartInfo['Size'] = self.DigikeyPartInfo.PartInfo['Package / Case']
        except:
            pass

        try:
            self.PartInfo['Temperature'] = self.DigikeyPartInfo.PartInfo['Operating Temperature']#[len('Temperature '):]
        except:
            pass

        #elapstedtime = time.time() - starttime
        #print("Parameter Details Updated in {:.3}s".format(elapstedtime))

    def ColumnOrderUpdate(self):
        current_column = self.Result_DataFrame.columns.values
        new_column = self.indexlist[:15]

        for n in range(len(current_column)):
            if current_column[n] == ("Compliance","RoHS"):
                new_column.append((current_column[n]))
        for n in range(len(current_column)):
            if current_column[n][0] == "Compliance" and current_column[n][1] != "RoHS" :
                new_column.append((current_column[n]))

        Price_Column_min = len(current_column)
        Price_Column_max = 0
        for n in range(len(current_column)): # Price column update
            if current_column[n][0] == "Mouser Price [CAD]":
                if n < Price_Column_min:
                    Price_Column_min = n
                if n > Price_Column_max:
                    Price_Column_max = n

        Price_Column_Order = {}
        for n in range(Price_Column_max-Price_Column_min+1):
            idx = current_column[n+Price_Column_min][1].find(':')+1
            try:
                if current_column[n+Price_Column_min][0] == 'Mouser Price [CAD]':
                    Price_Column_Order[n+Price_Column_min] = int(current_column[n+Price_Column_min][1][idx:])
            except:
                pass

        sorted_x = sorted(Price_Column_Order.items(), key=lambda kv: kv[1])

        for target_n_key, target_n_value in sorted_x:
            for n in range(len(current_column)):
                if n == target_n_key:
                    new_column.append(current_column[n])

        # for n in range(len(current_column)): # Price column update
        #     if current_column[n][0] == "Price [CAD]":
        #         new_column.append(current_column[n])

        new_column.append(("Data Sheet",""))
        new_column.append(("Part Image",""))

        self.Result_DataFrame = self.Result_DataFrame.reindex(columns=new_column)

    def AutoExcelWidthAdjuster(self, path="FilePath"):
        # read input xlsx
        wb1 = xl.load_workbook(filename=path)
        ws1 = wb1.worksheets[0]

        # set column width
        for col in ws1.columns:
            max_length = 0
            column = col[0].column

            for cell in col:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))

            adjusted_width = (max_length + 2)
            ws1.column_dimensions[col[1].column_letter].width = adjusted_width

        # border setting
        border = Border(top=Side(style='dotted', color='000000'),
                        bottom=Side(style='dotted', color='000000'),
                        left=Side(style='thin', color='000000'),
                        right=Side(style='thin', color='000000')
                        )

        for row_num in range(1, ws1.max_row+1):
            for col_num in range(1, ws1.max_column+1):
                ws1.cell(row=row_num, column=col_num).border = border
                ws1.cell(row=row_num, column=col_num).alignment = Alignment(horizontal='center',
                                                                            vertical='center',
                                                                            wrap_text=False)


        ws1.cell(row=1, column=1).value = 'Index'
        ws1.views.activeCell = 'A3'
        # save xlsx file
        wb1.save(path)

    def WritingExcel(self):
        print("\nMaking Result Sheet at Desktop...")
        textwriter("\nMaking Result Sheet at Desktop...")

        try:
            path = os.environ['DataAnalyzerPath'] + str('\\AnalysisTool\\DataAnalyzer\\assets\\')
        except:
            path = os.getcwd() + str('\\DataAnalyzer\\assets\\')  # For normal operation
            # if __name__ == "__main__":
            #     path = os.getcwd() + str('\\assets\\')      #For debuging case
            # else:
            #     path = os.getcwd() + str('\\DataAnalyzer\\assets\\')   #For normal operation
        path = r'{}'.format(path)

        path_desktop = os.environ['USERPROFILE'] + str('\\Desktop\\')
        try:
            os.system('TASKKILL /F /IM excel.exe')
        except:
            pass
        ### Delete old file
        try:
            os.remove(path_desktop + "DerivedExcel.xlsx")
        except:
            pass

        ### Copy all excel format and conditional format settings
        shutil.copyfile(path + 'Base.xlsx', path_desktop + 'DerivedExcel.xlsx')


        book = load_workbook(path_desktop + 'DerivedExcel.xlsx')
        writer = pd.ExcelWriter(path_desktop + 'DerivedExcel.xlsx', engine='openpyxl')
        writer.book = book

        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

        self.Result_DataFrame.to_excel(writer, "Sheet1",)
        writer.save()
        writer.close()

        self.AutoExcelWidthAdjuster(path_desktop + 'DerivedExcel.xlsx')
        self.DigikeyPartInfo.terminate()

        subprocess.Popen(['start', path_desktop + 'DerivedExcel.xlsx'], shell=True)

if __name__ == "__main__":

    apikey=os.environ['MOUSER_API']

    TEST =6
    if TEST == 1: # Part from text
            PartNumber = "ERJ-3GEYJ390V"
            Mouser = API_Mouser(PN=PartNumber, APIKEY=apikey)
            print(json.dumps(Mouser.PartInfo, indent=4)) #All derived data from Mouser
            print('Part Number : {}'.format(Mouser.PartParameter("ManufacturerPartNumber")))
            print('Description : {}'.format(Mouser.PartParameter("Description")))
            print('Voltage Rating : {}'.format(Mouser.PartParameter("VoltageRating")))
            print('Wattage Rating : {}'.format(Mouser.PartParameter("Rating")))
            print('Value : {}'.format(Mouser.PartParameter("Value")))
            print('Dielectric : {}'.format(Mouser.PartParameter("Dielectric")))
            print('Tolerance : {}'.format(Mouser.PartParameter("Tolerance")))
            print('Size : {}'.format(Mouser.PartParameter("Size")))
            print('RoHs : {}'.format(Mouser.PartParameter("ROHSStatus")))

    elif TEST == 2: #Parts from List
            PartNumberList = ["CHPHT0603K1002FGT",
                              "CHPHT0603LR100JGT",
                              "PEG228KKL4110QE1",
                              "05085C104K4T2A",
                              "05085C104K4T"    ##Dammy Intentional Wrong Part Number
                              ]
            MouserList = API_Mouser(PNList=PartNumberList,APIKEY=apikey)

    elif TEST == 3: #Parts from Excel File
        path = "C:\\Users\\iyoneda\\Desktop\\BOM_rA.xlsm"
        API_Mouser(APIKEY=apikey,PATH=path)
        print(path)

    elif TEST == 4:
        PartNumberList = ["LT1491AIS#PBF",
                            "CHPHT0603K1002FGT",
                          #"CHPHT0603LR100JGT",
                          #"PEG228KKL4110QE1",
                          "05085C104K4T2A",
                          ]
        MouserList = API_Mouser(PNList=PartNumberList, APIKEY=apikey)
        #MouserList.WritingExcel()

    elif TEST == 5:
        PartNumberList = ["BLM21PG600SN1D",
                          "MAX13450EAUD+",
                          ]
        MouserList = API_Mouser(PNList=PartNumberList, APIKEY=apikey)

    elif TEST == 6:
        PartList = ""
        List = API_Mouser(PNList=PartList, APIKEY=apikey)